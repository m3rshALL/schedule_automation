from celery import shared_task
from apps.academics.models import Course
from apps.faculty.models import Teacher
from apps.facilities.models import Room
from apps.schedule.models import TimeSlot, Schedule
from utils.optapy_transform import course_to_fact, teacher_to_fact, room_to_fact, timeslot_to_fact
import openpyxl
import os
from .models import TaskHistory
from django.contrib.auth import get_user_model
from openpyxl.utils import get_column_letter
import tempfile
from django.conf import settings

@shared_task
def optimize_schedule_task():
    """
    Celery-задача для запуска OptaPy-оптимизации расписания.
    """
    # 1. Получить все необходимые данные
    courses = list(Course.objects.all())
    teachers = list(Teacher.objects.all())
    rooms = list(Room.objects.all())
    timeslots = list(TimeSlot.objects.all())

    # 2. Преобразовать в OptaPy DTO
    course_facts = [course_to_fact(c) for c in courses]
    teacher_facts = [teacher_to_fact(t) for t in teachers]
    room_facts = [room_to_fact(r) for r in rooms]
    timeslot_facts = [timeslot_to_fact(ts) for ts in timeslots]

    # 3. Сформировать planning entities (ScheduleEntry) и solution
    # TODO: реализовать генерацию ScheduleEntry и запуск OptaPy solver
    # from apps.schedule.optapy_entities import ScheduleSolution
    # solution = ScheduleSolution(...)
    # result = solve_with_optapy(solution)

    # 4. Сохранить результат в Schedule
    # TODO: очистить старое расписание и записать новое
    # for entry in result.schedule_entries:
    #     Schedule.objects.create(...)

    return "Optimization started (mock)"

@shared_task(bind=True)
def import_schedule_task(self, file_path, user_id=None):
    """
    Импорт расписания из Excel-файла (формат как при экспорте).
    file_path: путь к .xlsx файлу
    user_id: id пользователя, инициировавшего импорт
    """
    imported = 0
    errors = []
    # Создаём/обновляем TaskHistory
    if user_id:
        TaskHistory.objects.update_or_create(
            task_id=self.request.id,
            defaults={
                'user_id': user_id,
                'type': 'import',
                'status': 'STARTED',
            }
        )
    try:
        wb = openpyxl.load_workbook(file_path)
        for ws in wb.worksheets:
            for row in ws.iter_rows(min_row=2, values_only=True):
                try:
                    # Пример: [Время, Группа, Кол-во студентов, Дисциплина, Преподаватель, Email, Аудитория, Блок, Оборудование, Тип занятия]
                    time_str, group_name, group_size, subject_name, teacher_name, email, room_name, block, equipment, lesson_type = row[:10]
                    day, time_range = ws.title, time_str
                    start_time, end_time = time_range.split('-')
                    # Найти или создать объекты
                    timeslot, _ = TimeSlot.objects.get_or_create(day=day, start_time=start_time.strip(), end_time=end_time.strip())
                    course = Course.objects.filter(subject__name=subject_name, student_group__name=group_name).first()
                    teacher = Teacher.objects.filter(name=teacher_name).first()
                    room = Room.objects.filter(name=room_name).first()
                    if not all([course, teacher, room, timeslot]):
                        errors.append(f"Не найдены объекты для строки: {row}")
                        continue
                    Schedule.objects.create(course=course, teacher=teacher, room=room, timeslot=timeslot)
                    imported += 1
                except Exception as e:
                    errors.append(str(e))
        wb.close()
        os.remove(file_path)
        status = 'SUCCESS'
    except Exception as e:
        errors.append(str(e))
        status = 'FAILURE'
    # Обновляем TaskHistory
    if user_id:
        TaskHistory.objects.update_or_create(
            task_id=self.request.id,
            defaults={
                'user_id': user_id,
                'type': 'import',
                'status': status,
                'result_data': {'imported': imported, 'errors': errors},
            }
        )
    return {"imported": imported, "errors": errors}

@shared_task(bind=True)
def export_schedule_task(self, user_id=None, export_format='excel'):
    status = 'STARTED'
    file_path = None
    try:
        wb = openpyxl.Workbook()
        wb.remove(wb.active)
        days = ['Monday', 'Tuesday', 'Wednesday', 'Thursday', 'Friday', 'Saturday']
        day_map = {
            'Monday': 'Понедельник',
            'Tuesday': 'Вторник',
            'Wednesday': 'Среда',
            'Thursday': 'Четверг',
            'Friday': 'Пятница',
            'Saturday': 'Суббота',
            'Sunday': 'Воскресенье',
        }
        schedules = Schedule.objects.select_related(
            'course', 'teacher', 'room', 'timeslot', 'course__subject', 'course__student_group'
        ).all().order_by('timeslot__day', 'timeslot__start_time')
        for day in days:
            ws = wb.create_sheet(title=day_map.get(day, day))
            headers = [
                'Время', 'Группа', 'Кол-во студентов', 'Дисциплина', 'Преподаватель', 'Email',
                'Аудитория', 'Блок', 'Оборудование', 'Тип занятия'
            ]
            ws.append(headers)
            for cell in ws[1]:
                cell.font = openpyxl.styles.Font(bold=True)
            day_schedules = [s for s in schedules if s.timeslot.day == day]
            for s in sorted(day_schedules, key=lambda x: x.timeslot.start_time):
                equipment = ', '.join([e.name for e in s.room.equipment.all()])
                ws.append([
                    f"{s.timeslot.start_time:%H:%M}-{s.timeslot.end_time:%H:%M}",
                    s.course.student_group.name,
                    s.course.student_group.size,
                    s.course.subject.name,
                    s.teacher.name,
                    getattr(s.teacher, 'email', ''),
                    s.room.name,
                    getattr(s.room, 'block', ''),
                    equipment,
                    s.course.lesson_type,
                ])
            ws.auto_filter.ref = ws.dimensions
            for col in range(1, len(headers) + 1):
                ws.column_dimensions[get_column_letter(col)].width = 18
        tmp_dir = getattr(settings, 'MEDIA_ROOT', tempfile.gettempdir())
        file_path = os.path.join(tmp_dir, f'schedule_export_{os.getpid()}.xlsx')
        wb.save(file_path)
        status = 'SUCCESS'
    except Exception as e:
        status = 'FAILURE'
        file_path = None
    # Обновляем TaskHistory
    if user_id:
        TaskHistory.objects.update_or_create(
            task_id=self.request.id,
            defaults={
                'user_id': user_id,
                'type': 'export',
                'status': status,
                'result_url': file_path if status == 'SUCCESS' else None,
            }
        )
    return file_path 