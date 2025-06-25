from django.shortcuts import render
from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework import status
from celery.result import AsyncResult
from .tasks import optimize_schedule_task, import_schedule_task
import openpyxl
from openpyxl.utils import get_column_letter
from django.http import HttpResponse
from .models import Schedule, TimeSlot, TaskHistory
from rest_framework import permissions
import io
from reportlab.lib.pagesizes import A4
from reportlab.pdfgen import canvas
from celery import shared_task
import os
from django.conf import settings
import tempfile
from rest_framework import viewsets, filters
from django_filters.rest_framework import DjangoFilterBackend
from .serializers import TimeSlotSerializer, ScheduleSerializer
from django.contrib.auth.decorators import login_required
from django.utils.decorators import method_decorator
from django.views import View
from django.shortcuts import redirect
from django.urls import reverse
from django.contrib import messages
from django.utils import timezone
from rest_framework.permissions import IsAuthenticated

# Create your views here.

class OptimizeScheduleView(APIView):
    """
    POST /api/schedule/optimize/ — запуск задачи оптимизации
    """
    def post(self, request):
        task = optimize_schedule_task.delay()
        return Response({"task_id": task.id}, status=status.HTTP_202_ACCEPTED)

class OptimizeScheduleStatusView(APIView):
    """
    GET /api/schedule/optimize/status/?task_id=... — статус задачи
    """
    def get(self, request):
        task_id = request.query_params.get("task_id")
        if not task_id:
            return Response({"error": "task_id required"}, status=status.HTTP_400_BAD_REQUEST)
        result = AsyncResult(task_id)
        return Response({
            "task_id": task_id,
            "status": result.status,
            "result": result.result if result.successful() else None
        })

class ScheduleExportView(APIView):
    permission_classes = [permissions.IsAdminUser]

    def get(self, request):
        export_format = request.query_params.get('format', 'excel')
        schedules = Schedule.objects.select_related(
            'course', 'teacher', 'room', 'timeslot', 'course__subject', 'course__student_group'
        ).all().order_by('timeslot__day', 'timeslot__start_time')
        if export_format == 'pdf':
            return self.export_pdf(schedules)
        return self.export_excel(schedules)

    def export_excel(self, schedules):
        wb = openpyxl.Workbook()
        wb.remove(wb.active)
        days = ['Monday', 'Tuesday', 'Wednesday', 'Thursday', 'Friday', 'Saturday']
        day_map = {
            'Monday': 'Понедельник',
            'Tuesday': 'Вторник',
            'Wednesday': 'Среда',
            'Thursday': 'Четверг',
            'Friday': 'Пятница',
            'Saturday': 'Суббота',
            'Sunday': 'Воскресенье',
        }
        for day in days:
            ws = wb.create_sheet(title=day_map.get(day, day))
            headers = [
                'Время', 'Группа', 'Кол-во студентов', 'Дисциплина', 'Преподаватель', 'Email',
                'Аудитория', 'Блок', 'Оборудование', 'Тип занятия'
            ]
            ws.append(headers)
            for cell in ws[1]:
                cell.font = openpyxl.styles.Font(bold=True)
            day_schedules = [s for s in schedules if s.timeslot.day == day]
            for s in sorted(day_schedules, key=lambda x: x.timeslot.start_time):
                equipment = ', '.join([e.name for e in s.room.equipment.all()])
                ws.append([
                    f"{s.timeslot.start_time:%H:%M}-{s.timeslot.end_time:%H:%M}",
                    s.course.student_group.name,
                    s.course.student_group.size,
                    s.course.subject.name,
                    s.teacher.name,
                    getattr(s.teacher, 'email', ''),
                    s.room.name,
                    getattr(s.room, 'block', ''),
                    equipment,
                    s.course.lesson_type,
                ])
            ws.auto_filter.ref = ws.dimensions
            for col in range(1, len(headers) + 1):
                ws.column_dimensions[get_column_letter(col)].width = 18
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        response = HttpResponse(
            output.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename=schedule_grouped_by_day.xlsx'
        return response

    def export_pdf(self, schedules):
        buffer = io.BytesIO()
        p = canvas.Canvas(buffer, pagesize=A4)
        width, height = A4
        y = height - 40
        p.setFont('Helvetica-Bold', 14)
        p.drawString(40, y, 'Расписание занятий')
        y -= 30
        p.setFont('Helvetica', 10)
        headers = ['Группа', 'Дисциплина', 'Преподаватель', 'Аудитория', 'День', 'Время', 'Тип']
        for i, h in enumerate(headers):
            p.drawString(40 + i*80, y, h)
        y -= 20
        for s in schedules:
            row = [
                s.course.student_group.name,
                s.course.subject.name,
                s.teacher.name,
                s.room.name,
                s.timeslot.day,
                f"{s.timeslot.start_time:%H:%M}-{s.timeslot.end_time:%H:%M}",
                s.course.lesson_type,
            ]
            for i, val in enumerate(row):
                p.drawString(40 + i*80, y, str(val))
            y -= 18
            if y < 60:
                p.showPage()
                y = height - 40
        p.save()
        buffer.seek(0)
        response = HttpResponse(buffer, content_type='application/pdf')
        response['Content-Disposition'] = 'attachment; filename=schedule.pdf'
        return response

@shared_task
def export_schedule_task():
    import openpyxl
    from openpyxl.utils import get_column_letter
    from apps.schedule.models import Schedule
    import tempfile
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    days = ['Monday', 'Tuesday', 'Wednesday', 'Thursday', 'Friday', 'Saturday']
    day_map = {
        'Monday': 'Понедельник',
        'Tuesday': 'Вторник',
        'Wednesday': 'Среда',
        'Thursday': 'Четверг',
        'Friday': 'Пятница',
        'Saturday': 'Суббота',
        'Sunday': 'Воскресенье',
    }
    schedules = Schedule.objects.select_related(
        'course', 'teacher', 'room', 'timeslot', 'course__subject', 'course__student_group'
    ).all().order_by('timeslot__day', 'timeslot__start_time')
    for day in days:
        ws = wb.create_sheet(title=day_map.get(day, day))
        headers = [
            'Время', 'Группа', 'Кол-во студентов', 'Дисциплина', 'Преподаватель', 'Email',
            'Аудитория', 'Блок', 'Оборудование', 'Тип занятия'
        ]
        ws.append(headers)
        for cell in ws[1]:
            cell.font = openpyxl.styles.Font(bold=True)
        day_schedules = [s for s in schedules if s.timeslot.day == day]
        for s in sorted(day_schedules, key=lambda x: x.timeslot.start_time):
            equipment = ', '.join([e.name for e in s.room.equipment.all()])
            ws.append([
                f"{s.timeslot.start_time:%H:%M}-{s.timeslot.end_time:%H:%M}",
                s.course.student_group.name,
                s.course.student_group.size,
                s.course.subject.name,
                s.teacher.name,
                getattr(s.teacher, 'email', ''),
                s.room.name,
                getattr(s.room, 'block', ''),
                equipment,
                s.course.lesson_type,
            ])
        ws.auto_filter.ref = ws.dimensions
        for col in range(1, len(headers) + 1):
            ws.column_dimensions[get_column_letter(col)].width = 18
    tmp_dir = getattr(settings, 'MEDIA_ROOT', tempfile.gettempdir())
    file_path = os.path.join(tmp_dir, f'schedule_export_{os.getpid()}.xlsx')
    wb.save(file_path)
    return file_path

class ScheduleAsyncExportView(APIView):
    permission_classes = [permissions.IsAdminUser]
    def post(self, request):
        task = export_schedule_task.delay()
        return Response({'task_id': task.id}, status=status.HTTP_202_ACCEPTED)

class ScheduleAsyncExportStatusView(APIView):
    permission_classes = [permissions.IsAdminUser]
    def get(self, request):
        task_id = request.query_params.get('task_id')
        if not task_id:
            return Response({'error': 'task_id required'}, status=status.HTTP_400_BAD_REQUEST)
        result = AsyncResult(task_id)
        if result.state == 'SUCCESS':
            file_path = result.result
            if not os.path.exists(file_path):
                return Response({'error': 'file not found'}, status=status.HTTP_404_NOT_FOUND)
            with open(file_path, 'rb') as f:
                response = HttpResponse(f.read(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
                response['Content-Disposition'] = 'attachment; filename=schedule_async.xlsx'
                return response
        return Response({'task_id': task_id, 'status': result.state}, status=status.HTTP_200_OK)

class TimeSlotViewSet(viewsets.ModelViewSet):
    queryset = TimeSlot.objects.all()
    serializer_class = TimeSlotSerializer
    filter_backends = [DjangoFilterBackend, filters.SearchFilter, filters.OrderingFilter]
    filterset_fields = ["day", "start_time", "end_time"]
    search_fields = ["day"]
    ordering_fields = ["day", "start_time", "end_time"]
    permission_classes = [permissions.IsAdminUser]

class ScheduleViewSet(viewsets.ModelViewSet):
    queryset = Schedule.objects.select_related("course", "teacher", "room", "timeslot").all()
    serializer_class = ScheduleSerializer
    filter_backends = [DjangoFilterBackend, filters.SearchFilter, filters.OrderingFilter]
    filterset_fields = ["course", "teacher", "room", "timeslot"]
    search_fields = ["course__subject__name", "teacher__name", "room__name"]
    ordering_fields = ["timeslot", "teacher", "room"]
    permission_classes = [permissions.IsAdminUser]

class ScheduleImportExportView(View):
    @method_decorator(login_required)
    def get(self, request):
        # История задач пользователя из базы
        tasks = TaskHistory.objects.filter(user=request.user).order_by('-created_at')[:20]
        return render(request, 'schedule_import_export.html', {'tasks': tasks})

    @method_decorator(login_required)
    def post(self, request):
        if 'import_file' in request.FILES:
            import_file = request.FILES['import_file']
            tmp_dir = tempfile.gettempdir()
            tmp_path = os.path.join(tmp_dir, f"import_{request.user.id}_{int(timezone.now().timestamp())}.xlsx")
            with open(tmp_path, 'wb+') as f:
                for chunk in import_file.chunks():
                    f.write(chunk)
            task = import_schedule_task.delay(tmp_path, user_id=request.user.id)
            TaskHistory.objects.create(
                user=request.user,
                task_id=task.id,
                type='import',
                status='PENDING',
            )
            messages.success(request, 'Импорт запущен!')
        elif request.POST.get('export') == '1' or request.GET.get('format'):
            export_format = request.POST.get('format') or request.GET.get('format') or 'excel'
            task = export_schedule_task.delay(user_id=request.user.id, export_format=export_format)
            TaskHistory.objects.create(
                user=request.user,
                task_id=task.id,
                type='export',
                status='PENDING',
            )
            messages.success(request, 'Экспорт запущен!')
        else:
            messages.error(request, 'Неизвестное действие')
        return redirect(reverse('schedule-import-export'))

class TaskHistoryStatusAPIView(APIView):
    permission_classes = [IsAuthenticated]
    def get(self, request):
        tasks = TaskHistory.objects.filter(user=request.user).order_by('-created_at')[:20]
        data = [
            {
                'task_id': t.task_id,
                'type': t.get_type_display(),
                'status': t.status,
                'created_at': t.created_at.isoformat(),
                'result_url': t.result_url,
                'result_data': t.result_data,
            }
            for t in tasks
        ]
        return Response({'tasks': data})
