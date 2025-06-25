from rest_framework import viewsets, permissions, filters
from django_filters.rest_framework import DjangoFilterBackend
import openpyxl
from openpyxl.utils import get_column_letter
from django.http import HttpResponse
import io
from reportlab.lib.pagesizes import A4
from reportlab.pdfgen import canvas
from rest_framework.views import APIView
from celery import shared_task
import os
from django.conf import settings
from celery.result import AsyncResult
from rest_framework.response import Response
from rest_framework import status

# Create your views here.

from .models import Course, StudentGroup, Subject
from .serializers import (
    CourseSerializer,
    StudentGroupSerializer,
    SubjectSerializer,
)


class SubjectViewSet(viewsets.ModelViewSet):
    queryset = Subject.objects.all()
    serializer_class = SubjectSerializer
    filter_backends = [DjangoFilterBackend, filters.SearchFilter, filters.OrderingFilter]
    filterset_fields = ["name", "code"]
    search_fields = ["name", "code"]
    ordering_fields = ["name", "code"]
    permission_classes = [permissions.IsAdminUser]


class StudentGroupViewSet(viewsets.ModelViewSet):
    queryset = StudentGroup.objects.all()
    serializer_class = StudentGroupSerializer
    filter_backends = [DjangoFilterBackend, filters.SearchFilter, filters.OrderingFilter]
    filterset_fields = ["name", "size"]
    search_fields = ["name"]
    ordering_fields = ["name", "size"]
    permission_classes = [permissions.IsAdminUser]


class CourseViewSet(viewsets.ModelViewSet):
    queryset = Course.objects.select_related("subject", "student_group").all()
    serializer_class = CourseSerializer
    filter_backends = [DjangoFilterBackend, filters.SearchFilter, filters.OrderingFilter]
    filterset_fields = [
        "subject",
        "student_group",
        "course_year",
        "lesson_type",
        "is_mook",
        "is_elective",
        "elective_group",
        "parent_course",
    ]
    search_fields = ["subject__name", "student_group__name"]
    ordering_fields = ["course_year", "lesson_type"]
    permission_classes = [permissions.IsAdminUser]


class StudentGroupExportView(APIView):
    permission_classes = [permissions.IsAdminUser]
    def get(self, request):
        export_format = request.query_params.get('format', 'excel')
        groups = StudentGroup.objects.all().order_by('name')
        if export_format == 'pdf':
            return self.export_pdf(groups)
        return self.export_excel(groups)
    def export_excel(self, groups):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Группы'
        headers = ['Название', 'Количество студентов']
        ws.append(headers)
        for cell in ws[1]:
            cell.font = openpyxl.styles.Font(bold=True)
        for g in groups:
            ws.append([g.name, g.size])
        ws.auto_filter.ref = ws.dimensions
        for col in range(1, len(headers) + 1):
            ws.column_dimensions[get_column_letter(col)].width = 22
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        response = HttpResponse(
            output.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename=groups.xlsx'
        return response
    def export_pdf(self, groups):
        buffer = io.BytesIO()
        p = canvas.Canvas(buffer, pagesize=A4)
        width, height = A4
        y = height - 40
        p.setFont('Helvetica-Bold', 14)
        p.drawString(40, y, 'Список групп студентов')
        y -= 30
        p.setFont('Helvetica', 10)
        headers = ['Название', 'Количество студентов']
        for i, h in enumerate(headers):
            p.drawString(40 + i*180, y, h)
        y -= 20
        for g in groups:
            row = [g.name, g.size]
            for i, val in enumerate(row):
                p.drawString(40 + i*180, y, str(val))
            y -= 18
            if y < 60:
                p.showPage()
                y = height - 40
        p.save()
        buffer.seek(0)
        response = HttpResponse(buffer, content_type='application/pdf')
        response['Content-Disposition'] = 'attachment; filename=groups.pdf'
        return response

class CourseExportView(APIView):
    permission_classes = [permissions.IsAdminUser]
    def get(self, request):
        export_format = request.query_params.get('format', 'excel')
        courses = Course.objects.select_related('subject', 'student_group').all().order_by('subject__name', 'student_group__name')
        if export_format == 'pdf':
            return self.export_pdf(courses)
        return self.export_excel(courses)
    def export_excel(self, courses):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Курсы'
        headers = [
            'Дисциплина', 'Группа', 'Часы в неделю', 'Год', 'Тип занятия', 'MOOK',
            'Электив', 'Группа электива', 'Родительский курс', 'Требуемое оборудование'
        ]
        ws.append(headers)
        for cell in ws[1]:
            cell.font = openpyxl.styles.Font(bold=True)
        for c in courses:
            req_eq = ', '.join([e.name for e in c.required_equipment.all()])
            ws.append([
                c.subject.name,
                c.student_group.name,
                c.hours_per_week,
                c.course_year,
                c.lesson_type,
                'Да' if c.is_mook else 'Нет',
                'Да' if c.is_elective else 'Нет',
                c.elective_group or '',
                c.parent_course.subject.name if c.parent_course else '',
                req_eq
            ])
        ws.auto_filter.ref = ws.dimensions
        for col in range(1, len(headers) + 1):
            ws.column_dimensions[get_column_letter(col)].width = 22
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        response = HttpResponse(
            output.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename=courses.xlsx'
        return response
    def export_pdf(self, courses):
        buffer = io.BytesIO()
        p = canvas.Canvas(buffer, pagesize=A4)
        width, height = A4
        y = height - 40
        p.setFont('Helvetica-Bold', 14)
        p.drawString(40, y, 'Список курсов')
        y -= 30
        p.setFont('Helvetica', 8)
        headers = [
            'Дисциплина', 'Группа', 'Часы/нед', 'Год', 'Тип', 'MOOK',
            'Электив', 'Гр. электива', 'Род. курс', 'Оборудование'
        ]
        for i, h in enumerate(headers):
            p.drawString(40 + i*70, y, h)
        y -= 20
        for c in courses:
            req_eq = ', '.join([e.name for e in c.required_equipment.all()])
            row = [
                c.subject.name,
                c.student_group.name,
                c.hours_per_week,
                c.course_year,
                c.lesson_type,
                'Да' if c.is_mook else 'Нет',
                'Да' if c.is_elective else 'Нет',
                c.elective_group or '',
                c.parent_course.subject.name if c.parent_course else '',
                req_eq
            ]
            for i, val in enumerate(row):
                p.drawString(40 + i*70, y, str(val))
            y -= 18
            if y < 60:
                p.showPage()
                y = height - 40
        p.save()
        buffer.seek(0)
        response = HttpResponse(buffer, content_type='application/pdf')
        response['Content-Disposition'] = 'attachment; filename=courses.pdf'
        return response

@shared_task
def export_groups_task():
    import openpyxl
    from openpyxl.utils import get_column_letter
    from apps.academics.models import StudentGroup
    import io
    import tempfile
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Группы'
    headers = ['Название', 'Количество студентов']
    ws.append(headers)
    for cell in ws[1]:
        cell.font = openpyxl.styles.Font(bold=True)
    for g in StudentGroup.objects.all().order_by('name'):
        ws.append([g.name, g.size])
    ws.auto_filter.ref = ws.dimensions
    for col in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 22
    tmp_dir = getattr(settings, 'MEDIA_ROOT', tempfile.gettempdir())
    file_path = os.path.join(tmp_dir, f'groups_export_{os.getpid()}.xlsx')
    wb.save(file_path)
    return file_path

class StudentGroupAsyncExportView(APIView):
    permission_classes = [permissions.IsAdminUser]
    def post(self, request):
        task = export_groups_task.delay()
        return Response({'task_id': task.id}, status=status.HTTP_202_ACCEPTED)

class StudentGroupAsyncExportStatusView(APIView):
    permission_classes = [permissions.IsAdminUser]
    def get(self, request):
        task_id = request.query_params.get('task_id')
        if not task_id:
            return Response({'error': 'task_id required'}, status=status.HTTP_400_BAD_REQUEST)
        result = AsyncResult(task_id)
        if result.state == 'SUCCESS':
            file_path = result.result
            if not os.path.exists(file_path):
                return Response({'error': 'file not found'}, status=status.HTTP_404_NOT_FOUND)
            with open(file_path, 'rb') as f:
                response = HttpResponse(f.read(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
                response['Content-Disposition'] = 'attachment; filename=groups_async.xlsx'
                return response
        return Response({'task_id': task_id, 'status': result.state}, status=status.HTTP_200_OK)

@shared_task
def export_courses_task():
    import openpyxl
    from openpyxl.utils import get_column_letter
    from apps.academics.models import Course
    import tempfile
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Курсы'
    headers = [
        'Дисциплина', 'Группа', 'Часы в неделю', 'Год', 'Тип занятия', 'MOOK',
        'Электив', 'Группа электива', 'Родительский курс', 'Требуемое оборудование'
    ]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = openpyxl.styles.Font(bold=True)
    for c in Course.objects.select_related('subject', 'student_group').all().order_by('subject__name', 'student_group__name'):
        req_eq = ', '.join([e.name for e in c.required_equipment.all()])
        ws.append([
            c.subject.name,
            c.student_group.name,
            c.hours_per_week,
            c.course_year,
            c.lesson_type,
            'Да' if c.is_mook else 'Нет',
            'Да' if c.is_elective else 'Нет',
            c.elective_group or '',
            c.parent_course.subject.name if c.parent_course else '',
            req_eq
        ])
    ws.auto_filter.ref = ws.dimensions
    for col in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 22
    tmp_dir = getattr(settings, 'MEDIA_ROOT', tempfile.gettempdir())
    file_path = os.path.join(tmp_dir, f'courses_export_{os.getpid()}.xlsx')
    wb.save(file_path)
    return file_path

class CourseAsyncExportView(APIView):
    permission_classes = [permissions.IsAdminUser]
    def post(self, request):
        task = export_courses_task.delay()
        return Response({'task_id': task.id}, status=status.HTTP_202_ACCEPTED)

class CourseAsyncExportStatusView(APIView):
    permission_classes = [permissions.IsAdminUser]
    def get(self, request):
        task_id = request.query_params.get('task_id')
        if not task_id:
            return Response({'error': 'task_id required'}, status=status.HTTP_400_BAD_REQUEST)
        result = AsyncResult(task_id)
        if result.state == 'SUCCESS':
            file_path = result.result
            if not os.path.exists(file_path):
                return Response({'error': 'file not found'}, status=status.HTTP_404_NOT_FOUND)
            with open(file_path, 'rb') as f:
                response = HttpResponse(f.read(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
                response['Content-Disposition'] = 'attachment; filename=courses_async.xlsx'
                return response
        return Response({'task_id': task_id, 'status': result.state}, status=status.HTTP_200_OK)
