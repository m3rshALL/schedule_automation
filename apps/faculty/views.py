from rest_framework import viewsets, permissions, filters
from django_filters.rest_framework import DjangoFilterBackend
import openpyxl
from openpyxl.utils import get_column_letter
from django.http import HttpResponse
import io
from reportlab.lib.pagesizes import A4
from reportlab.pdfgen import canvas
from rest_framework.views import APIView
from celery import shared_task
import os
from django.conf import settings
from celery.result import AsyncResult
from rest_framework.response import Response
from rest_framework import status

# Create your views here.

from .models import Teacher
from .serializers import TeacherSerializer


class TeacherViewSet(viewsets.ModelViewSet):
    queryset = Teacher.objects.prefetch_related("qualifications").all()
    serializer_class = TeacherSerializer
    filter_backends = [DjangoFilterBackend, filters.SearchFilter, filters.OrderingFilter]
    filterset_fields = ["name", "qualifications"]
    search_fields = ["name", "qualifications__name"]
    ordering_fields = ["name"]
    permission_classes = [permissions.IsAdminUser]


class TeacherExportView(APIView):
    permission_classes = [permissions.IsAdminUser]
    def get(self, request):
        export_format = request.query_params.get('format', 'excel')
        teachers = Teacher.objects.prefetch_related('qualifications').all().order_by('name')
        if export_format == 'pdf':
            return self.export_pdf(teachers)
        return self.export_excel(teachers)
    def export_excel(self, teachers):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Преподаватели'
        headers = ['ФИО', 'Email', 'Квалификации (дисциплины)']
        ws.append(headers)
        for cell in ws[1]:
            cell.font = openpyxl.styles.Font(bold=True)
        for t in teachers:
            qualifications = ', '.join([s.name for s in t.qualifications.all()])
            ws.append([t.name, getattr(t, 'email', ''), qualifications])
        ws.auto_filter.ref = ws.dimensions
        for col in range(1, len(headers) + 1):
            ws.column_dimensions[get_column_letter(col)].width = 30
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        response = HttpResponse(
            output.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename=teachers.xlsx'
        return response
    def export_pdf(self, teachers):
        buffer = io.BytesIO()
        p = canvas.Canvas(buffer, pagesize=A4)
        width, height = A4
        y = height - 40
        p.setFont('Helvetica-Bold', 14)
        p.drawString(40, y, 'Список преподавателей')
        y -= 30
        p.setFont('Helvetica', 10)
        headers = ['ФИО', 'Email', 'Квалификации (дисциплины)']
        for i, h in enumerate(headers):
            p.drawString(40 + i*180, y, h)
        y -= 20
        for t in teachers:
            qualifications = ', '.join([s.name for s in t.qualifications.all()])
            row = [t.name, getattr(t, 'email', ''), qualifications]
            for i, val in enumerate(row):
                p.drawString(40 + i*180, y, str(val))
            y -= 18
            if y < 60:
                p.showPage()
                y = height - 40
        p.save()
        buffer.seek(0)
        response = HttpResponse(buffer, content_type='application/pdf')
        response['Content-Disposition'] = 'attachment; filename=teachers.pdf'
        return response

@shared_task
def export_teachers_task():
    import openpyxl
    from openpyxl.utils import get_column_letter
    from apps.faculty.models import Teacher
    import tempfile
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Преподаватели'
    headers = ['ФИО', 'Email', 'Квалификации (дисциплины)']
    ws.append(headers)
    for cell in ws[1]:
        cell.font = openpyxl.styles.Font(bold=True)
    for t in Teacher.objects.prefetch_related('qualifications').all().order_by('name'):
        qualifications = ', '.join([s.name for s in t.qualifications.all()])
        ws.append([t.name, getattr(t, 'email', ''), qualifications])
    ws.auto_filter.ref = ws.dimensions
    for col in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 30
    tmp_dir = getattr(settings, 'MEDIA_ROOT', tempfile.gettempdir())
    file_path = os.path.join(tmp_dir, f'teachers_export_{os.getpid()}.xlsx')
    wb.save(file_path)
    return file_path

class TeacherAsyncExportView(APIView):
    permission_classes = [permissions.IsAdminUser]
    def post(self, request):
        task = export_teachers_task.delay()
        return Response({'task_id': task.id}, status=status.HTTP_202_ACCEPTED)

class TeacherAsyncExportStatusView(APIView):
    permission_classes = [permissions.IsAdminUser]
    def get(self, request):
        task_id = request.query_params.get('task_id')
        if not task_id:
            return Response({'error': 'task_id required'}, status=status.HTTP_400_BAD_REQUEST)
        result = AsyncResult(task_id)
        if result.state == 'SUCCESS':
            file_path = result.result
            if not os.path.exists(file_path):
                return Response({'error': 'file not found'}, status=status.HTTP_404_NOT_FOUND)
            with open(file_path, 'rb') as f:
                response = HttpResponse(f.read(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
                response['Content-Disposition'] = 'attachment; filename=teachers_async.xlsx'
                return response
        return Response({'task_id': task_id, 'status': result.state}, status=status.HTTP_200_OK)
