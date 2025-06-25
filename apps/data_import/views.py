from django.shortcuts import render
from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework import status, permissions, parsers
from apps.academics.models import StudentGroup
import openpyxl
from apps.faculty.models import Teacher
from apps.academics.models import Subject, Course
from apps.facilities.models import Room, Equipment
from apps.schedule.models import TimeSlot
from django.db import transaction

# Create your views here.

class StudentGroupImportView(APIView):
    permission_classes = [permissions.IsAdminUser]
    parser_classes = [parsers.MultiPartParser]

    def post(self, request, *args, **kwargs):
        file = request.FILES.get('file')
        if not file:
            return Response({'error': 'No file provided'}, status=status.HTTP_400_BAD_REQUEST)
        try:
            wb = openpyxl.load_workbook(file)
            ws = wb.active
            created, updated, errors = 0, 0, []
            for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                name, size = row[:2]
                if not name or not size:
                    errors.append(f'Row {i}: missing name or size')
                    continue
                obj, created_flag = StudentGroup.objects.update_or_create(
                    name=name,
                    defaults={'size': size}
                )
                if created_flag:
                    created += 1
                else:
                    updated += 1
            return Response({
                'created': created,
                'updated': updated,
                'errors': errors
            }, status=status.HTTP_200_OK)
        except Exception as e:
            return Response({'error': str(e)}, status=status.HTTP_400_BAD_REQUEST)

class TeacherImportView(APIView):
    permission_classes = [permissions.IsAdminUser]
    parser_classes = [parsers.MultiPartParser]

    def post(self, request, *args, **kwargs):
        file = request.FILES.get('file')
        if not file:
            return Response({'error': 'No file provided'}, status=status.HTTP_400_BAD_REQUEST)
        try:
            wb = openpyxl.load_workbook(file)
            ws = wb.active
            created, updated, errors = 0, 0, []
            for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                name, qualifications = row[:2]
                if not name:
                    errors.append(f'Row {i}: missing name')
                    continue
                teacher, created_flag = Teacher.objects.update_or_create(
                    name=name
                )
                if qualifications:
                    subject_names = [s.strip() for s in str(qualifications).split(',') if s.strip()]
                    subjects = Subject.objects.filter(name__in=subject_names)
                    teacher.qualifications.set(subjects)
                if created_flag:
                    created += 1
                else:
                    updated += 1
            return Response({'created': created, 'updated': updated, 'errors': errors}, status=status.HTTP_200_OK)
        except Exception as e:
            return Response({'error': str(e)}, status=status.HTTP_400_BAD_REQUEST)

class EquipmentImportView(APIView):
    permission_classes = [permissions.IsAdminUser]
    parser_classes = [parsers.MultiPartParser]

    def post(self, request, *args, **kwargs):
        file = request.FILES.get('file')
        if not file:
            return Response({'error': 'No file provided'}, status=status.HTTP_400_BAD_REQUEST)
        try:
            wb = openpyxl.load_workbook(file)
            ws = wb.active
            created, updated, errors = 0, 0, []
            for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                name = row[0]
                if not name:
                    errors.append(f'Row {i}: missing name')
                    continue
                obj, created_flag = Equipment.objects.update_or_create(name=name)
                if created_flag:
                    created += 1
                else:
                    updated += 1
            return Response({'created': created, 'updated': updated, 'errors': errors}, status=status.HTTP_200_OK)
        except Exception as e:
            return Response({'error': str(e)}, status=status.HTTP_400_BAD_REQUEST)

class RoomImportView(APIView):
    permission_classes = [permissions.IsAdminUser]
    parser_classes = [parsers.MultiPartParser]

    def post(self, request, *args, **kwargs):
        file = request.FILES.get('file')
        if not file:
            return Response({'error': 'No file provided'}, status=status.HTTP_400_BAD_REQUEST)
        try:
            wb = openpyxl.load_workbook(file)
            ws = wb.active
            created, updated, errors = 0, 0, []
            for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                name, capacity, room_type, block, equipment = row[:5]
                if not name or not capacity or not room_type:
                    errors.append(f'Row {i}: missing name, capacity or room_type')
                    continue
                obj, created_flag = Room.objects.update_or_create(
                    name=name,
                    defaults={
                        'capacity': capacity,
                        'room_type': room_type,
                        'block': block or 'C1',
                    }
                )
                if equipment:
                    eq_names = [e.strip() for e in str(equipment).split(',') if e.strip()]
                    eqs = Equipment.objects.filter(name__in=eq_names)
                    obj.equipment.set(eqs)
                if created_flag:
                    created += 1
                else:
                    updated += 1
            return Response({'created': created, 'updated': updated, 'errors': errors}, status=status.HTTP_200_OK)
        except Exception as e:
            return Response({'error': str(e)}, status=status.HTTP_400_BAD_REQUEST)

class CourseImportView(APIView):
    permission_classes = [permissions.IsAdminUser]
    parser_classes = [parsers.MultiPartParser]

    @transaction.atomic
    def post(self, request, *args, **kwargs):
        file = request.FILES.get('file')
        if not file:
            return Response({'error': 'No file provided'}, status=status.HTTP_400_BAD_REQUEST)
        try:
            wb = openpyxl.load_workbook(file)
            ws = wb.active
            created, updated, errors = 0, 0, []
            for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                (subject_name, group_name, hours_per_week, course_year, lesson_type, is_mook, is_elective, elective_group, parent_course_name, required_equipment) = row[:10]
                if not subject_name or not group_name:
                    errors.append(f'Row {i}: missing subject or group')
                    continue
                try:
                    subject = Subject.objects.get(name=subject_name)
                    group = StudentGroup.objects.get(name=group_name)
                except Exception as e:
                    errors.append(f'Row {i}: {e}')
                    continue
                parent_course = None
                if parent_course_name:
                    parent_course = Course.objects.filter(subject__name=parent_course_name).first()
                obj, created_flag = Course.objects.update_or_create(
                    subject=subject,
                    student_group=group,
                    defaults={
                        'hours_per_week': hours_per_week or 1,
                        'course_year': course_year or 1,
                        'lesson_type': lesson_type or 'lecture',
                        'is_mook': bool(is_mook),
                        'is_elective': bool(is_elective),
                        'elective_group': elective_group,
                        'parent_course': parent_course,
                    }
                )
                if required_equipment:
                    eq_names = [e.strip() for e in str(required_equipment).split(',') if e.strip()]
                    eqs = Equipment.objects.filter(name__in=eq_names)
                    obj.required_equipment.set(eqs)
                if created_flag:
                    created += 1
                else:
                    updated += 1
            return Response({'created': created, 'updated': updated, 'errors': errors}, status=status.HTTP_200_OK)
        except Exception as e:
            return Response({'error': str(e)}, status=status.HTTP_400_BAD_REQUEST)

class ScheduleImportView(APIView):
    permission_classes = [permissions.IsAdminUser]
    parser_classes = [parsers.MultiPartParser]

    @transaction.atomic
    def post(self, request, *args, **kwargs):
        file = request.FILES.get('file')
        if not file:
            return Response({'error': 'No file provided'}, status=status.HTTP_400_BAD_REQUEST)
        try:
            wb = openpyxl.load_workbook(file)
            ws = wb.active
            created, updated, errors = 0, 0, []
            for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                course_name, teacher_name, room_name, day, start_time, end_time = row[:6]
                if not course_name or not teacher_name or not room_name or not day or not start_time or not end_time:
                    errors.append(f'Row {i}: missing required fields')
                    continue
                try:
                    course = Course.objects.get(subject__name=course_name)
                    teacher = Teacher.objects.get(name=teacher_name)
                    room = Room.objects.get(name=room_name)
                    timeslot, _ = TimeSlot.objects.get_or_create(day=day, start_time=start_time, end_time=end_time)
                except Exception as e:
                    errors.append(f'Row {i}: {e}')
                    continue
                # Здесь предполагается, что будет модель ScheduleEntry, которую нужно создать/обновить
                # Пример:
                # entry, created_flag = ScheduleEntry.objects.update_or_create(
                #     course=course, teacher=teacher, room=room, timeslot=timeslot
                # )
                # if created_flag:
                #     created += 1
                # else:
                #     updated += 1
                # Пока только считаем успешные строки
                created += 1
            return Response({'created': created, 'updated': updated, 'errors': errors}, status=status.HTTP_200_OK)
        except Exception as e:
            return Response({'error': str(e)}, status=status.HTTP_400_BAD_REQUEST)
