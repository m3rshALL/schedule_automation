from rest_framework import viewsets, permissions, filters
from django_filters.rest_framework import DjangoFilterBackend
import openpyxl
from openpyxl.utils import get_column_letter
from django.http import HttpResponse
import io
from reportlab.lib.pagesizes import A4
from reportlab.pdfgen import canvas
from celery import shared_task
import os
from django.conf import settings
from celery.result import AsyncResult
from rest_framework.response import Response
from rest_framework import status

# Create your views here.

from .models import Equipment, Room
from .serializers import EquipmentSerializer, RoomSerializer

from rest_framework.views import APIView

class EquipmentViewSet(viewsets.ModelViewSet):
    """
    API endpoint that allows equipment to be viewed or edited.
    """

    queryset = Equipment.objects.all()
    serializer_class = EquipmentSerializer
    filter_backends = [DjangoFilterBackend, filters.SearchFilter, filters.OrderingFilter]
    filterset_fields = ["name"]
    search_fields = ["name"]
    ordering_fields = ["name"]
    permission_classes = [permissions.IsAdminUser]


class RoomViewSet(viewsets.ModelViewSet):
    """
    API endpoint that allows rooms to be viewed or edited.
    """

    queryset = Room.objects.all().prefetch_related("equipment")
    serializer_class = RoomSerializer
    filter_backends = [DjangoFilterBackend, filters.SearchFilter, filters.OrderingFilter]
    filterset_fields = ["name", "capacity", "room_type", "block", "equipment"]
    search_fields = ["name", "block"]
    ordering_fields = ["name", "capacity", "room_type", "block"]
    permission_classes = [permissions.IsAdminUser]

class RoomExportView(APIView):
    permission_classes = [permissions.IsAdminUser]
    def get(self, request):
        export_format = request.query_params.get('format', 'excel')
        rooms = Room.objects.prefetch_related('equipment').all().order_by('name')
        if export_format == 'pdf':
            return self.export_pdf(rooms)
        return self.export_excel(rooms)
    def export_excel(self, rooms):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Аудитории'
        headers = ['Название', 'Вместимость', 'Тип', 'Блок', 'Оборудование']
        ws.append(headers)
        for cell in ws[1]:
            cell.font = openpyxl.styles.Font(bold=True)
        for r in rooms:
            eq = ', '.join([e.name for e in r.equipment.all()])
            ws.append([r.name, r.capacity, r.get_room_type_display(), r.block, eq])
        ws.auto_filter.ref = ws.dimensions
        for col in range(1, len(headers) + 1):
            ws.column_dimensions[get_column_letter(col)].width = 18
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        response = HttpResponse(
            output.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename=rooms.xlsx'
        return response
    def export_pdf(self, rooms):
        buffer = io.BytesIO()
        p = canvas.Canvas(buffer, pagesize=A4)
        width, height = A4
        y = height - 40
        p.setFont('Helvetica-Bold', 14)
        p.drawString(40, y, 'Список аудиторий')
        y -= 30
        p.setFont('Helvetica', 10)
        headers = ['Название', 'Вместимость', 'Тип', 'Блок', 'Оборудование']
        for i, h in enumerate(headers):
            p.drawString(40 + i*80, y, h)
        y -= 20
        for r in rooms:
            eq = ', '.join([e.name for e in r.equipment.all()])
            row = [r.name, r.capacity, r.get_room_type_display(), r.block, eq]
            for i, val in enumerate(row):
                p.drawString(40 + i*80, y, str(val))
            y -= 18
            if y < 60:
                p.showPage()
                y = height - 40
        p.save()
        buffer.seek(0)
        response = HttpResponse(buffer, content_type='application/pdf')
        response['Content-Disposition'] = 'attachment; filename=rooms.pdf'
        return response

class EquipmentExportView(APIView):
    permission_classes = [permissions.IsAdminUser]
    def get(self, request):
        export_format = request.query_params.get('format', 'excel')
        equipment = Equipment.objects.all().order_by('name')
        if export_format == 'pdf':
            return self.export_pdf(equipment)
        return self.export_excel(equipment)
    def export_excel(self, equipment):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Оборудование'
        headers = ['Название']
        ws.append(headers)
        for cell in ws[1]:
            cell.font = openpyxl.styles.Font(bold=True)
        for e in equipment:
            ws.append([e.name])
        ws.auto_filter.ref = ws.dimensions
        ws.column_dimensions[get_column_letter(1)].width = 30
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        response = HttpResponse(
            output.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename=equipment.xlsx'
        return response
    def export_pdf(self, equipment):
        buffer = io.BytesIO()
        p = canvas.Canvas(buffer, pagesize=A4)
        width, height = A4
        y = height - 40
        p.setFont('Helvetica-Bold', 14)
        p.drawString(40, y, 'Список оборудования')
        y -= 30
        p.setFont('Helvetica', 10)
        p.drawString(40, y, 'Название')
        y -= 20
        for e in equipment:
            p.drawString(40, y, e.name)
            y -= 18
            if y < 60:
                p.showPage()
                y = height - 40
        p.save()
        buffer.seek(0)
        response = HttpResponse(buffer, content_type='application/pdf')
        response['Content-Disposition'] = 'attachment; filename=equipment.pdf'
        return response

@shared_task
def export_rooms_task():
    import openpyxl
    from openpyxl.utils import get_column_letter
    from apps.facilities.models import Room
    import tempfile
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Аудитории'
    headers = ['Название', 'Вместимость', 'Тип', 'Блок', 'Оборудование']
    ws.append(headers)
    for cell in ws[1]:
        cell.font = openpyxl.styles.Font(bold=True)
    for r in Room.objects.prefetch_related('equipment').all().order_by('name'):
        eq = ', '.join([e.name for e in r.equipment.all()])
        ws.append([r.name, r.capacity, r.get_room_type_display(), r.block, eq])
    ws.auto_filter.ref = ws.dimensions
    for col in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 18
    tmp_dir = getattr(settings, 'MEDIA_ROOT', tempfile.gettempdir())
    file_path = os.path.join(tmp_dir, f'rooms_export_{os.getpid()}.xlsx')
    wb.save(file_path)
    return file_path

class RoomAsyncExportView(APIView):
    permission_classes = [permissions.IsAdminUser]
    def post(self, request):
        task = export_rooms_task.delay()
        return Response({'task_id': task.id}, status=status.HTTP_202_ACCEPTED)

class RoomAsyncExportStatusView(APIView):
    permission_classes = [permissions.IsAdminUser]
    def get(self, request):
        task_id = request.query_params.get('task_id')
        if not task_id:
            return Response({'error': 'task_id required'}, status=status.HTTP_400_BAD_REQUEST)
        result = AsyncResult(task_id)
        if result.state == 'SUCCESS':
            file_path = result.result
            if not os.path.exists(file_path):
                return Response({'error': 'file not found'}, status=status.HTTP_404_NOT_FOUND)
            with open(file_path, 'rb') as f:
                response = HttpResponse(f.read(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
                response['Content-Disposition'] = 'attachment; filename=rooms_async.xlsx'
                return response
        return Response({'task_id': task_id, 'status': result.state}, status=status.HTTP_200_OK)

@shared_task
def export_equipment_task():
    import openpyxl
    from openpyxl.utils import get_column_letter
    from apps.facilities.models import Equipment
    import tempfile
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Оборудование'
    headers = ['Название']
    ws.append(headers)
    for cell in ws[1]:
        cell.font = openpyxl.styles.Font(bold=True)
    for e in Equipment.objects.all().order_by('name'):
        ws.append([e.name])
    ws.auto_filter.ref = ws.dimensions
    ws.column_dimensions[get_column_letter(1)].width = 30
    tmp_dir = getattr(settings, 'MEDIA_ROOT', tempfile.gettempdir())
    file_path = os.path.join(tmp_dir, f'equipment_export_{os.getpid()}.xlsx')
    wb.save(file_path)
    return file_path

class EquipmentAsyncExportView(APIView):
    permission_classes = [permissions.IsAdminUser]
    def post(self, request):
        task = export_equipment_task.delay()
        return Response({'task_id': task.id}, status=status.HTTP_202_ACCEPTED)

class EquipmentAsyncExportStatusView(APIView):
    permission_classes = [permissions.IsAdminUser]
    def get(self, request):
        task_id = request.query_params.get('task_id')
        if not task_id:
            return Response({'error': 'task_id required'}, status=status.HTTP_400_BAD_REQUEST)
        result = AsyncResult(task_id)
        if result.state == 'SUCCESS':
            file_path = result.result
            if not os.path.exists(file_path):
                return Response({'error': 'file not found'}, status=status.HTTP_404_NOT_FOUND)
            with open(file_path, 'rb') as f:
                response = HttpResponse(f.read(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
                response['Content-Disposition'] = 'attachment; filename=equipment_async.xlsx'
                return response
        return Response({'task_id': task_id, 'status': result.state}, status=status.HTTP_200_OK)
